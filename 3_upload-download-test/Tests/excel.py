import openpyxl

class Excel_upd():
    def __init__(self, excel_file_path):
        # Открываем файл
        self.file_path = excel_file_path
        self.book = openpyxl.load_workbook(self.file_path)  # Сохраняем книгу в self.book
        self.sheet = self.book.active
        self.column_index = None
        self.target_cell = None

    def read_exact_name(self, fruit_name, column_name):
        # Находим индекс столбца "Price"
        for j in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(row=1, column=j).value == column_name:
                self.column_index = j
                break
        # Ищем фрукт по имени
        for i in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=i, column=2).value == fruit_name:
                self.target_cell=self.sheet.cell(row=i, column=self.column_index)
                return self.target_cell

    def add_new_value(self, new_value):
        self.target_cell.value=new_value
        self.book.save(self.file_path)
        return self.target_cell.value


